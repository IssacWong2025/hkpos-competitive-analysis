from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median
from typing import Dict, List

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
MR_DATA_DIR = ROOT / "market_research" / "data"

COMPETITORS_CSV = DATA_DIR / "competitors_master.csv"
META_CSV = DATA_DIR / "meta_ads_intel.csv"
SEMRUSH_CSV = DATA_DIR / "semrush_google_ads_signals.csv"
EVIDENCE_V2_XLSX = MR_DATA_DIR / "evidence_log_v2_20260303.xlsx"

OUT_PANEL_CSV = MR_DATA_DIR / "competitor_signal_panel_v2_20260303.csv"
OUT_MODEL_INPUTS_CSV = MR_DATA_DIR / "model_inputs_v4_seed_20260303.csv"
OUT_SNAPSHOT_JSON = ROOT / "docs" / "data" / "market_share_v2_snapshot.json"

# Practical V2 gate (can be tightened after more S-tier collection):
# pass if S>=2 OR (S>=1 and A_effective>=2)
MIN_S_STRICT = 2
MIN_S_WITH_A = 1
MIN_A_EFFECTIVE = 2


def read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def norm_comp(name: str) -> str:
    return (name or "").strip().lower()


def read_evidence_xlsx(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows: List[Dict[str, str]] = []
    for i in range(2, ws.max_row + 1):
        item: Dict[str, str] = {}
        has = False
        for j, h in enumerate(headers, start=1):
            v = ws.cell(i, j).value
            s = "" if v is None else str(v).strip()
            item[h] = s
            if s:
                has = True
        if has:
            rows.append(item)
    return rows


def to_num(v: str) -> float | None:
    try:
        if v is None:
            return None
        s = str(v).strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def safe_int(v: str) -> int:
    n = to_num(v)
    if n is None:
        return 0
    return int(n)


def mode_or_default(values: List[str], default: str = "unknown") -> str:
    vals = [v for v in values if v]
    if not vals:
        return default
    return Counter(vals).most_common(1)[0][0]


def main() -> None:
    competitors = read_csv(COMPETITORS_CSV)
    meta_rows = read_csv(META_CSV)
    semrush_rows = read_csv(SEMRUSH_CSV)
    ev_rows = read_evidence_xlsx(EVIDENCE_V2_XLSX)

    by_comp_meta: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for r in meta_rows:
        by_comp_meta[norm_comp(r.get("competitor_name", ""))].append(r)

    by_domain_semrush: Dict[str, Dict[str, str]] = {}
    for r in semrush_rows:
        d = (r.get("website_domain", "") or "").strip().lower()
        if d and d not in by_domain_semrush:
            by_domain_semrush[d] = r

    ev_counts: Dict[str, Counter] = defaultdict(Counter)
    for r in ev_rows:
        comp = norm_comp(r.get("competitor_name", ""))
        tier = (r.get("evidence_tier", "") or "").strip().upper()
        status = (r.get("verification_status", "") or "").strip().lower()
        if not comp or tier not in {"S", "A", "B"}:
            continue
        if status in {"rejected", "invalid"}:
            continue
        ev_counts[comp][tier] += 1

    panel_rows: List[Dict[str, str]] = []
    v4_seed_rows: List[Dict[str, str]] = []

    for c in competitors:
        comp = (c.get("competitor_name", "") or "").strip()
        comp_key = norm_comp(comp)
        domain = (c.get("website_domain", "") or "").strip().lower()
        metas = by_comp_meta.get(comp_key, [])
        sem = by_domain_semrush.get(domain, {})

        ad_counts = [safe_int(r.get("ad_count_active", "")) for r in metas]
        active_ads = max(ad_counts) if ad_counts else 0
        obj_primary = mode_or_default([r.get("objective_path_hint", "") for r in metas], "unknown")
        meta_active_flag = any((r.get("status", "") or "").lower() == "active" for r in metas)

        paid_kw_count = safe_int(sem.get("paid_keywords_count", ""))
        semrush_has_data = paid_kw_count > 0

        s_cnt = int(ev_counts.get(comp_key, Counter()).get("S", 0))
        a_cnt = int(ev_counts.get(comp_key, Counter()).get("A", 0))
        b_cnt = int(ev_counts.get(comp_key, Counter()).get("B", 0))

        # A-tier includes growth intelligence from Meta/Semrush as required by V2 plan.
        a_effective = a_cnt + (1 if meta_active_flag else 0) + (1 if semrush_has_data else 0)
        sufficient = (s_cnt >= MIN_S_STRICT) or (s_cnt >= MIN_S_WITH_A and a_effective >= MIN_A_EFFECTIVE)

        row = {
            "snapshot_date": "2026-03-03",
            "competitor_name": comp,
            "website_domain": domain,
            "s_evidence_count": str(s_cnt),
            "a_evidence_count": str(a_cnt),
            "b_evidence_count": str(b_cnt),
            "a_effective_count": str(a_effective),
            "meta_active_flag": "Y" if meta_active_flag else "N",
            "meta_ad_count_active": str(active_ads),
            "meta_objective_primary": obj_primary,
            "semrush_paid_keywords_count": str(paid_kw_count),
            "semrush_has_data": "Y" if semrush_has_data else "N",
            "meets_min_evidence_gate": "Y" if sufficient else "N",
            "gate_rule": f"S>={MIN_S_STRICT} OR (S>={MIN_S_WITH_A} AND A_effective>={MIN_A_EFFECTIVE})",
        }
        panel_rows.append(row)

    # Seed file for V4 model: keep it simple and transparent for manual review.
    ad_values = [safe_int(r["meta_ad_count_active"]) for r in panel_rows]
    kw_values = [safe_int(r["semrush_paid_keywords_count"]) for r in panel_rows]
    med_ads = median(ad_values) if ad_values else 0
    med_kw = median(kw_values) if kw_values else 0
    max_ads = max(ad_values) if ad_values else 0
    max_kw = max(kw_values) if kw_values else 0

    for r in panel_rows:
        ads = safe_int(r["meta_ad_count_active"])
        kws = safe_int(r["semrush_paid_keywords_count"])
        s = to_num(r["s_evidence_count"]) or 0.0
        a_eff = to_num(r["a_effective_count"]) or 0.0
        b = to_num(r["b_evidence_count"]) or 0.0

        signal_s = s / 4.0 if s > 0 else 0.0
        signal_a = ((ads / max_ads) if max_ads else 0.0) * 0.5 + ((kws / max_kw) if max_kw else 0.0) * 0.5
        signal_b = min(b / 4.0, 1.0)
        conf = "A" if r["meets_min_evidence_gate"] == "Y" and s >= 2 else ("B" if r["meets_min_evidence_gate"] == "Y" else "C")

        v4_seed_rows.append(
            {
                "snapshot_date": r["snapshot_date"],
                "competitor_name": r["competitor_name"],
                "website_domain": r["website_domain"],
                "signal_s_direct": f"{signal_s:.4f}",
                "signal_a_proxy": f"{signal_a:.4f}",
                "signal_b_proxy": f"{signal_b:.4f}",
                "weight_s": "0.60",
                "weight_a": "0.30",
                "weight_b": "0.10",
                "confidence_level": conf,
                "meets_min_evidence_gate": r["meets_min_evidence_gate"],
                "notes": f"ads_median={med_ads},kw_median={med_kw},a_effective={int(a_eff)}",
            }
        )

    panel_headers = [
        "snapshot_date",
        "competitor_name",
        "website_domain",
        "s_evidence_count",
        "a_evidence_count",
        "b_evidence_count",
        "a_effective_count",
        "meta_active_flag",
        "meta_ad_count_active",
        "meta_objective_primary",
        "semrush_paid_keywords_count",
        "semrush_has_data",
        "meets_min_evidence_gate",
        "gate_rule",
    ]
    OUT_PANEL_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PANEL_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=panel_headers)
        w.writeheader()
        for r in panel_rows:
            w.writerow(r)

    seed_headers = [
        "snapshot_date",
        "competitor_name",
        "website_domain",
        "signal_s_direct",
        "signal_a_proxy",
        "signal_b_proxy",
        "weight_s",
        "weight_a",
        "weight_b",
        "confidence_level",
        "meets_min_evidence_gate",
        "notes",
    ]
    with OUT_MODEL_INPUTS_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=seed_headers)
        w.writeheader()
        for r in v4_seed_rows:
            w.writerow(r)

    passed = sum(1 for r in panel_rows if r["meets_min_evidence_gate"] == "Y")
    snapshot = {
        "snapshot_date": "2026-03-03",
        "gate_rule": f"S>={MIN_S_STRICT} OR (S>={MIN_S_WITH_A} AND A_effective>={MIN_A_EFFECTIVE})",
        "pass_count": passed,
        "total_count": len(panel_rows),
        "competitors": panel_rows,
    }
    OUT_SNAPSHOT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_SNAPSHOT_JSON.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Panel generated: {OUT_PANEL_CSV}")
    print(f"V4 seed generated: {OUT_MODEL_INPUTS_CSV}")
    print(f"Snapshot generated: {OUT_SNAPSHOT_JSON}")
    print(f"Evidence gate pass: {passed}/{len(panel_rows)}")


if __name__ == "__main__":
    main()
